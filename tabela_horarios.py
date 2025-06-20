import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

def gerar_tabelas_por_periodo(disciplinas, xlsx_path):
    """
    Gera uma aba de horários para cada período (campo 'periodo') no arquivo Excel.
    disciplinas: lista de dicts com 'nome', 'horario', 'periodo', ...
    xlsx_path: caminho do arquivo de saída
    """
    dias_semana = ['Seg', 'Ter', 'Qua', 'Qui', 'Sex', 'Sáb']
    periodos = sorted(set(d['periodo'] for d in disciplinas))
    horarios = sorted(set(d['horario'].split(' ', 1)[1] for d in disciplinas))

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Estilos
    header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))
    center = Alignment(horizontal='center', vertical='center')
    bold = Font(bold=True)

    for periodo in periodos:
        ws = wb.create_sheet(title=f"Período {periodo}")

        # Cabeçalho
        ws.cell(row=1, column=1, value="Horários").font = bold
        ws.cell(row=1, column=1).alignment = center
        ws.cell(row=1, column=1).fill = header_fill
        ws.cell(row=1, column=1).border = border

        for col, dia in enumerate(dias_semana, start=2):
            cell = ws.cell(row=1, column=col, value=dia)
            cell.font = bold
            cell.alignment = center
            cell.fill = header_fill
            cell.border = border

        # Tabela de horários
        for row, horario in enumerate(horarios, start=2):
            ws.cell(row=row, column=1, value=horario).font = bold
            ws.cell(row=row, column=1).alignment = center
            ws.cell(row=row, column=1).fill = header_fill
            ws.cell(row=row, column=1).border = border

            for col, dia in enumerate(dias_semana, start=2):
                cell = ws.cell(row=row, column=col)
                cell.border = border
                cell.alignment = Alignment(wrap_text=True, vertical='top')

        # Preenche as disciplinas
        for d in disciplinas:
            if d["periodo"] != periodo:
                continue
            dia, hora = d['horario'].split(' ', 1)
            row = horarios.index(hora) + 2
            col = dias_semana.index(dia) + 2
            nome = d["nome"]
            prof = d["professor"]
            ws.cell(row=row, column=col, value=f"{nome}\n{prof}")

        # Ajuste de largura automática
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        length = len(str(cell.value))
                        if length > max_length:
                            max_length = length
                except Exception:
                    pass
            ws.column_dimensions[column].width = max_length + 2

    wb.save(xlsx_path)